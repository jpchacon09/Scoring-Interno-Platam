#!/usr/bin/env python3
"""
Crea Excel Interactivo tipo "Simulador de Scoring"

Permite ajustar parámetros y ver en tiempo real cómo cambian los scores:
- Pesos de componentes (Payment Performance, Plan, Deterioration)
- Penalizaciones por mora
- Bonificaciones por buen comportamiento
- Umbrales de categorización
- Pesos del sistema híbrido

El Excel tiene fórmulas que se recalculan automáticamente al cambiar parámetros.

Usage:
    python scripts/11_create_interactive_simulator.py
"""

import pandas as pd
import numpy as np
from pathlib import Path
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation

logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

# Paths
BASE_DIR = Path(__file__).parent.parent
PROCESSED_DIR = BASE_DIR / 'data' / 'processed'
HYBRID_FILE = PROCESSED_DIR / 'hybrid_scores.csv'
OUTPUT_FILE = BASE_DIR / 'SIMULADOR_SCORING_INTERACTIVO.xlsx'

def create_interactive_simulator():
    """Crea Excel interactivo con simulador de scoring"""

    logger.info("="*80)
    logger.info("CREANDO SIMULADOR DE SCORING INTERACTIVO")
    logger.info("="*80)

    # Cargar datos
    logger.info(f"\n1. Cargando datos...")
    df = pd.read_csv(HYBRID_FILE)
    logger.info(f"   ✓ Cargados {len(df):,} clientes")

    # Limitar a primeros 100 clientes para que sea más rápido
    df_sample = df.head(100).copy()
    logger.info(f"   • Usando muestra de {len(df_sample)} clientes para simulación")

    # Crear workbook
    logger.info(f"\n2. Creando Excel con fórmulas interactivas...")
    wb = Workbook()

    # ========================================================================
    # HOJA 1: PANEL DE CONTROL
    # ========================================================================
    logger.info("   • Hoja 1: Panel de Control")
    ws_control = wb.active
    ws_control.title = "🎛️ PANEL DE CONTROL"

    # Estilos
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(bold=True, size=12, color="FFFFFF")
    section_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    section_font = Font(bold=True, size=11, color="FFFFFF")
    param_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
    value_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    value_font = Font(bold=True, size=11, color="FFFFFF")

    # Título
    ws_control['A1'] = "SIMULADOR DE SCORING PLATAM V2.0 + HÍBRIDO"
    ws_control['A1'].font = Font(bold=True, size=16, color="2C3E50")
    ws_control['A1'].alignment = Alignment(horizontal='center')
    ws_control.merge_cells('A1:D1')

    ws_control['A2'] = "Ajusta los parámetros abajo y los scores se recalcularán automáticamente"
    ws_control['A2'].font = Font(italic=True, size=10, color="7F8C8D")
    ws_control['A2'].alignment = Alignment(horizontal='center')
    ws_control.merge_cells('A2:D2')

    row = 4

    # SECCIÓN 1: PESOS DE COMPONENTES V2.0
    ws_control[f'A{row}'] = "📊 PESOS DE COMPONENTES PLATAM V2.0"
    ws_control[f'A{row}'].fill = section_fill
    ws_control[f'A{row}'].font = section_font
    ws_control.merge_cells(f'A{row}:D{row}')
    row += 1

    ws_control[f'A{row}'] = "Parámetro"
    ws_control[f'B{row}'] = "Valor Actual"
    ws_control[f'C{row}'] = "Máximo"
    ws_control[f'D{row}'] = "% del Total"
    for col in ['A', 'B', 'C', 'D']:
        ws_control[f'{col}{row}'].fill = param_fill
        ws_control[f'{col}{row}'].font = Font(bold=True)
    row += 1

    # Payment Performance
    ws_control[f'A{row}'] = "Payment Performance (pts)"
    ws_control[f'B{row}'] = 600
    ws_control[f'B{row}'].fill = value_fill
    ws_control[f'B{row}'].font = value_font
    ws_control[f'C{row}'] = 1000
    ws_control[f'D{row}'] = f"=B{row}/1000"
    ws_control[f'D{row}'].number_format = '0%'
    payment_perf_cell = f'B{row}'
    row += 1

    # Payment Plan History
    ws_control[f'A{row}'] = "Payment Plan History (pts)"
    ws_control[f'B{row}'] = 150
    ws_control[f'B{row}'].fill = value_fill
    ws_control[f'B{row}'].font = value_font
    ws_control[f'C{row}'] = 1000
    ws_control[f'D{row}'] = f"=B{row}/1000"
    ws_control[f'D{row}'].number_format = '0%'
    payment_plan_cell = f'B{row}'
    row += 1

    # Deterioration Velocity
    ws_control[f'A{row}'] = "Deterioration Velocity (pts)"
    ws_control[f'B{row}'] = 250
    ws_control[f'B{row}'].fill = value_fill
    ws_control[f'B{row}'].font = value_font
    ws_control[f'C{row}'] = 1000
    ws_control[f'D{row}'] = f"=B{row}/1000"
    ws_control[f'D{row}'].number_format = '0%'
    deterioration_cell = f'B{row}'
    row += 1

    # Total
    ws_control[f'A{row}'] = "TOTAL (debe sumar 1000)"
    ws_control[f'A{row}'].font = Font(bold=True)
    ws_control[f'B{row}'] = f"={payment_perf_cell}+{payment_plan_cell}+{deterioration_cell}"
    ws_control[f'B{row}'].font = Font(bold=True, color="E74C3C")
    ws_control[f'D{row}'] = f"=B{row}/1000"
    ws_control[f'D{row}'].number_format = '0%'
    ws_control[f'D{row}'].font = Font(bold=True)
    total_components_cell = f'B{row}'
    row += 2

    # SECCIÓN 2: PENALIZACIONES Y BONIFICACIONES
    ws_control[f'A{row}'] = "⚡ PENALIZACIONES Y BONIFICACIONES"
    ws_control[f'A{row}'].fill = section_fill
    ws_control[f'A{row}'].font = section_font
    ws_control.merge_cells(f'A{row}:D{row}')
    row += 1

    ws_control[f'A{row}'] = "Concepto"
    ws_control[f'B{row}'] = "Valor"
    ws_control[f'C{row}'] = "Unidad"
    ws_control[f'D{row}'] = "Descripción"
    for col in ['A', 'B', 'C', 'D']:
        ws_control[f'{col}{row}'].fill = param_fill
        ws_control[f'{col}{row}'].font = Font(bold=True)
    row += 1

    # Penalización por DPD alto
    ws_control[f'A{row}'] = "Penalización DPD >30 días"
    ws_control[f'B{row}'] = 50
    ws_control[f'B{row}'].fill = value_fill
    ws_control[f'B{row}'].font = value_font
    ws_control[f'C{row}'] = "puntos"
    ws_control[f'D{row}'] = "Resta del score total"
    penalty_dpd_cell = f'B{row}'
    row += 1

    # Bonificación por pagos a tiempo
    ws_control[f'A{row}'] = "Bonificación >90% pagos a tiempo"
    ws_control[f'B{row}'] = 30
    ws_control[f'B{row}'].fill = value_fill
    ws_control[f'B{row}'].font = value_font
    ws_control[f'C{row}'] = "puntos"
    ws_control[f'D{row}'] = "Suma al score total"
    bonus_ontime_cell = f'B{row}'
    row += 1

    # Penalización por alta utilización
    ws_control[f'A{row}'] = "Penalización utilización >80%"
    ws_control[f'B{row}'] = 20
    ws_control[f'B{row}'].fill = value_fill
    ws_control[f'B{row}'].font = value_font
    ws_control[f'C{row}'] = "puntos"
    ws_control[f'D{row}'] = "Resta del score total"
    penalty_util_cell = f'B{row}'
    row += 1

    # Bonificación por cliente maduro
    ws_control[f'A{row}'] = "Bonificación cliente >24 meses"
    ws_control[f'B{row}'] = 25
    ws_control[f'B{row}'].fill = value_fill
    ws_control[f'B{row}'].font = value_font
    ws_control[f'C{row}'] = "puntos"
    ws_control[f'D{row}'] = "Suma al score total"
    bonus_mature_cell = f'B{row}'
    row += 2

    # SECCIÓN 3: PESOS DEL SISTEMA HÍBRIDO
    ws_control[f'A{row}'] = "🔄 PESOS DEL SISTEMA HÍBRIDO"
    ws_control[f'A{row}'].fill = section_fill
    ws_control[f'A{row}'].font = section_font
    ws_control.merge_cells(f'A{row}:D{row}')
    row += 1

    ws_control[f'A{row}'] = "Categoría Cliente"
    ws_control[f'B{row}'] = "Peso PLATAM"
    ws_control[f'C{row}'] = "Peso HCPN"
    ws_control[f'D{row}'] = "Total"
    for col in ['A', 'B', 'C', 'D']:
        ws_control[f'{col}{row}'].fill = param_fill
        ws_control[f'{col}{row}'].font = Font(bold=True)
    row += 1

    # Muy nuevo
    ws_control[f'A{row}'] = "Muy Nuevo (< 3 meses)"
    ws_control[f'B{row}'] = 0.30
    ws_control[f'B{row}'].fill = value_fill
    ws_control[f'B{row}'].font = value_font
    ws_control[f'B{row}'].number_format = '0%'
    ws_control[f'C{row}'] = f"=1-B{row}"
    ws_control[f'C{row}'].number_format = '0%'
    ws_control[f'D{row}'] = f"=B{row}+C{row}"
    ws_control[f'D{row}'].number_format = '0%'
    peso_muy_nuevo_cell = f'B{row}'
    row += 1

    # Nuevo
    ws_control[f'A{row}'] = "Nuevo (3-6 meses)"
    ws_control[f'B{row}'] = 0.40
    ws_control[f'B{row}'].fill = value_fill
    ws_control[f'B{row}'].font = value_font
    ws_control[f'B{row}'].number_format = '0%'
    ws_control[f'C{row}'] = f"=1-B{row}"
    ws_control[f'C{row}'].number_format = '0%'
    ws_control[f'D{row}'] = f"=B{row}+C{row}"
    ws_control[f'D{row}'].number_format = '0%'
    peso_nuevo_cell = f'B{row}'
    row += 1

    # Intermedio
    ws_control[f'A{row}'] = "Intermedio (6-12 meses)"
    ws_control[f'B{row}'] = 0.50
    ws_control[f'B{row}'].fill = value_fill
    ws_control[f'B{row}'].font = value_font
    ws_control[f'B{row}'].number_format = '0%'
    ws_control[f'C{row}'] = f"=1-B{row}"
    ws_control[f'C{row}'].number_format = '0%'
    ws_control[f'D{row}'] = f"=B{row}+C{row}"
    ws_control[f'D{row}'].number_format = '0%'
    peso_intermedio_cell = f'B{row}'
    row += 1

    # Establecido
    ws_control[f'A{row}'] = "Establecido (12-24 meses)"
    ws_control[f'B{row}'] = 0.60
    ws_control[f'B{row}'].fill = value_fill
    ws_control[f'B{row}'].font = value_font
    ws_control[f'B{row}'].number_format = '0%'
    ws_control[f'C{row}'] = f"=1-B{row}"
    ws_control[f'C{row}'].number_format = '0%'
    ws_control[f'D{row}'] = f"=B{row}+C{row}"
    ws_control[f'D{row}'].number_format = '0%'
    peso_establecido_cell = f'B{row}'
    row += 1

    # Maduro
    ws_control[f'A{row}'] = "Maduro (> 24 meses)"
    ws_control[f'B{row}'] = 0.70
    ws_control[f'B{row}'].fill = value_fill
    ws_control[f'B{row}'].font = value_font
    ws_control[f'B{row}'].number_format = '0%'
    ws_control[f'C{row}'] = f"=1-B{row}"
    ws_control[f'C{row}'].number_format = '0%'
    ws_control[f'D{row}'] = f"=B{row}+C{row}"
    ws_control[f'D{row}'].number_format = '0%'
    peso_maduro_cell = f'B{row}'
    row += 2

    # SECCIÓN 4: UMBRALES DE CATEGORIZACIÓN
    ws_control[f'A{row}'] = "📏 UMBRALES DE CATEGORIZACIÓN"
    ws_control[f'A{row}'].fill = section_fill
    ws_control[f'A{row}'].font = section_font
    ws_control.merge_cells(f'A{row}:D{row}')
    row += 1

    ws_control[f'A{row}'] = "Rating"
    ws_control[f'B{row}'] = "Score Mínimo"
    ws_control[f'C{row}'] = "Score Máximo"
    ws_control[f'D{row}'] = "Descripción"
    for col in ['A', 'B', 'C', 'D']:
        ws_control[f'{col}{row}'].fill = param_fill
        ws_control[f'{col}{row}'].font = Font(bold=True)
    row += 1

    ratings = [
        ('A+', 900, 1000, 'Excelente'),
        ('A', 850, 899, 'Muy Bueno'),
        ('A-', 800, 849, 'Bueno Superior'),
        ('B+', 750, 799, 'Bueno'),
        ('B', 700, 749, 'Aceptable Superior'),
        ('B-', 650, 699, 'Aceptable'),
        ('C+', 600, 649, 'Regular Superior'),
        ('C', 550, 599, 'Regular'),
        ('C-', 500, 549, 'Regular Bajo'),
        ('D', 400, 499, 'Deficiente'),
        ('F', 0, 399, 'Muy Deficiente')
    ]

    rating_cells = {}
    for rating, min_score, max_score, desc in ratings:
        ws_control[f'A{row}'] = rating
        ws_control[f'B{row}'] = min_score
        ws_control[f'B{row}'].fill = value_fill
        ws_control[f'B{row}'].font = value_font
        ws_control[f'C{row}'] = max_score
        ws_control[f'D{row}'] = desc
        rating_cells[rating] = (f'B{row}', f'C{row}')
        row += 1

    # Ajustar anchos de columna
    ws_control.column_dimensions['A'].width = 35
    ws_control.column_dimensions['B'].width = 15
    ws_control.column_dimensions['C'].width = 15
    ws_control.column_dimensions['D'].width = 30

    # Instrucciones
    row += 1
    ws_control[f'A{row}'] = "💡 INSTRUCCIONES:"
    ws_control[f'A{row}'].font = Font(bold=True, size=11, color="27AE60")
    ws_control.merge_cells(f'A{row}:D{row}')
    row += 1

    instructions = [
        "1. Cambia los valores en las celdas AZULES",
        "2. Los pesos de componentes deben sumar 1000 puntos",
        "3. Los scores se recalcularán automáticamente en la hoja 'Scores Simulados'",
        "4. Compara los resultados en la hoja 'Comparación'",
        "5. Usa la hoja 'Análisis' para ver distribuciones"
    ]

    for instruction in instructions:
        ws_control[f'A{row}'] = instruction
        ws_control[f'A{row}'].font = Font(italic=True, size=9)
        ws_control.merge_cells(f'A{row}:D{row}')
        row += 1

    # ========================================================================
    # HOJA 2: DATOS ORIGINALES
    # ========================================================================
    logger.info("   • Hoja 2: Datos Originales")
    ws_data = wb.create_sheet("📊 Datos Originales")

    # Seleccionar columnas clave
    data_cols = [
        'cedula', 'client_name', 'score_payment_performance',
        'score_payment_plan', 'score_deterioration', 'platam_score',
        'platam_rating', 'experian_score_normalized', 'hybrid_score',
        'hybrid_rating', 'payment_id_count', 'days_past_due_mean',
        'pct_late', 'pct_utilization', 'months_as_client',
        'categoria_madurez', 'peso_platam_usado'
    ]

    # Filtrar solo columnas que existan
    available_cols = [col for col in data_cols if col in df_sample.columns]
    df_export = df_sample[available_cols].copy()

    # Escribir headers
    for col_idx, col_name in enumerate(df_export.columns, 1):
        cell = ws_data.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Escribir datos
    for row_idx, row_data in enumerate(df_export.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws_data.cell(row=row_idx, column=col_idx, value=value)

    # ========================================================================
    # HOJA 3: SCORES SIMULADOS (con fórmulas)
    # ========================================================================
    logger.info("   • Hoja 3: Scores Simulados")
    ws_sim = wb.create_sheet("🎯 Scores Simulados")

    # Headers
    sim_headers = [
        'ID Cliente', 'Nombre', 'Comp 1 (Base)', 'Comp 2 (Base)', 'Comp 3 (Base)',
        'Comp 1 (Simulado)', 'Comp 2 (Simulado)', 'Comp 3 (Simulado)',
        'Score Base', 'Penalizaciones', 'Bonificaciones', 'Score Simulado PLATAM',
        'Score HCPN', 'Peso PLATAM', 'Peso HCPN', 'Score Híbrido Simulado',
        'Rating Simulado', 'Cambio vs Base'
    ]

    for col_idx, header in enumerate(sim_headers, 1):
        cell = ws_sim.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Fórmulas para cada cliente
    for row_idx in range(2, len(df_sample) + 2):
        data_row = row_idx  # Misma fila en hoja de datos

        # ID y Nombre (referencia a datos)
        ws_sim.cell(row=row_idx, column=1, value=f"='📊 Datos Originales'!A{data_row}")
        ws_sim.cell(row=row_idx, column=2, value=f"='📊 Datos Originales'!B{data_row}")

        # Componentes base (referencia a datos)
        ws_sim.cell(row=row_idx, column=3, value=f"='📊 Datos Originales'!C{data_row}")  # payment_perf
        ws_sim.cell(row=row_idx, column=4, value=f"='📊 Datos Originales'!D{data_row}")  # payment_plan
        ws_sim.cell(row=row_idx, column=5, value=f"='📊 Datos Originales'!E{data_row}")  # deterioration

        # Componentes simulados (con nuevos pesos)
        # Fórmula: (componente_base / peso_original) * peso_nuevo
        ws_sim.cell(row=row_idx, column=6, value=f"=(C{row_idx}/600)*'🎛️ PANEL DE CONTROL'!{payment_perf_cell}")
        ws_sim.cell(row=row_idx, column=7, value=f"=(D{row_idx}/150)*'🎛️ PANEL DE CONTROL'!{payment_plan_cell}")
        ws_sim.cell(row=row_idx, column=8, value=f"=(E{row_idx}/250)*'🎛️ PANEL DE CONTROL'!{deterioration_cell}")

        # Score base (suma de componentes simulados)
        ws_sim.cell(row=row_idx, column=9, value=f"=F{row_idx}+G{row_idx}+H{row_idx}")

        # Penalizaciones (DPD >30 + utilización >80%)
        dpd_penalty = f"=IF('📊 Datos Originales'!L{data_row}>30, '🎛️ PANEL DE CONTROL'!{penalty_dpd_cell}, 0)"
        util_penalty = f"=IF('📊 Datos Originales'!N{data_row}>80, '🎛️ PANEL DE CONTROL'!{penalty_util_cell}, 0)"
        ws_sim.cell(row=row_idx, column=10, value=f"={dpd_penalty}+{util_penalty}")

        # Bonificaciones (>90% a tiempo + >24 meses)
        ontime_bonus = f"=IF('📊 Datos Originales'!M{data_row}<10, '🎛️ PANEL DE CONTROL'!{bonus_ontime_cell}, 0)"
        mature_bonus = f"=IF('📊 Datos Originales'!O{data_row}>24, '🎛️ PANEL DE CONTROL'!{bonus_mature_cell}, 0)"
        ws_sim.cell(row=row_idx, column=11, value=f"={ontime_bonus}+{mature_bonus}")

        # Score PLATAM simulado (base - penalizaciones + bonificaciones)
        ws_sim.cell(row=row_idx, column=12, value=f"=MAX(0, MIN(1000, I{row_idx}-J{row_idx}+K{row_idx}))")

        # Score HCPN (referencia a datos)
        ws_sim.cell(row=row_idx, column=13, value=f"='📊 Datos Originales'!H{data_row}")

        # Peso PLATAM (dinámico según categoría de madurez)
        categoria_cell = f"'📊 Datos Originales'!P{data_row}"
        peso_formula = f"""=IF({categoria_cell}="muy_nuevo", '🎛️ PANEL DE CONTROL'!{peso_muy_nuevo_cell},
                            IF({categoria_cell}="nuevo", '🎛️ PANEL DE CONTROL'!{peso_nuevo_cell},
                            IF({categoria_cell}="intermedio", '🎛️ PANEL DE CONTROL'!{peso_intermedio_cell},
                            IF({categoria_cell}="establecido", '🎛️ PANEL DE CONTROL'!{peso_establecido_cell},
                            '🎛️ PANEL DE CONTROL'!{peso_maduro_cell}))))"""
        ws_sim.cell(row=row_idx, column=14, value=peso_formula)
        ws_sim.cell(row=row_idx, column=14).number_format = '0%'

        # Peso HCPN (1 - peso PLATAM)
        ws_sim.cell(row=row_idx, column=15, value=f"=1-N{row_idx}")
        ws_sim.cell(row=row_idx, column=15).number_format = '0%'

        # Score Híbrido Simulado
        ws_sim.cell(row=row_idx, column=16, value=f"=(L{row_idx}*N{row_idx})+(M{row_idx}*O{row_idx})")

        # Rating simulado (usando umbrales del panel de control)
        rating_formula = f"""=IF(P{row_idx}>='🎛️ PANEL DE CONTROL'!{rating_cells['A+'][0]}, "A+",
                            IF(P{row_idx}>='🎛️ PANEL DE CONTROL'!{rating_cells['A'][0]}, "A",
                            IF(P{row_idx}>='🎛️ PANEL DE CONTROL'!{rating_cells['A-'][0]}, "A-",
                            IF(P{row_idx}>='🎛️ PANEL DE CONTROL'!{rating_cells['B+'][0]}, "B+",
                            IF(P{row_idx}>='🎛️ PANEL DE CONTROL'!{rating_cells['B'][0]}, "B",
                            IF(P{row_idx}>='🎛️ PANEL DE CONTROL'!{rating_cells['B-'][0]}, "B-",
                            IF(P{row_idx}>='🎛️ PANEL DE CONTROL'!{rating_cells['C+'][0]}, "C+",
                            IF(P{row_idx}>='🎛️ PANEL DE CONTROL'!{rating_cells['C'][0]}, "C",
                            IF(P{row_idx}>='🎛️ PANEL DE CONTROL'!{rating_cells['C-'][0]}, "C-",
                            IF(P{row_idx}>='🎛️ PANEL DE CONTROL'!{rating_cells['D'][0]}, "D", "F"))))))))))"""
        ws_sim.cell(row=row_idx, column=17, value=rating_formula)

        # Cambio vs base
        ws_sim.cell(row=row_idx, column=18, value=f"=P{row_idx}-'📊 Datos Originales'!I{data_row}")

    # Ajustar anchos
    for col in range(1, 19):
        ws_sim.column_dimensions[chr(64 + col)].width = 15

    # ========================================================================
    # HOJA 4: COMPARACIÓN
    # ========================================================================
    logger.info("   • Hoja 4: Comparación")
    ws_comp = wb.create_sheet("📈 Comparación")

    # Estadísticas comparativas
    ws_comp['A1'] = "COMPARACIÓN: SCORES BASE vs SIMULADOS"
    ws_comp['A1'].font = Font(bold=True, size=14)
    ws_comp.merge_cells('A1:D1')

    stats_row = 3
    ws_comp[f'A{stats_row}'] = "Métrica"
    ws_comp[f'B{stats_row}'] = "Score Base"
    ws_comp[f'C{stats_row}'] = "Score Simulado"
    ws_comp[f'D{stats_row}'] = "Diferencia"
    for col in ['A', 'B', 'C', 'D']:
        ws_comp[f'{col}{stats_row}'].fill = header_fill
        ws_comp[f'{col}{stats_row}'].font = header_font
    stats_row += 1

    # Promedio
    ws_comp[f'A{stats_row}'] = "Promedio"
    ws_comp[f'B{stats_row}'] = f"=AVERAGE('📊 Datos Originales'!I:I)"
    ws_comp[f'C{stats_row}'] = f"=AVERAGE('🎯 Scores Simulados'!P:P)"
    ws_comp[f'D{stats_row}'] = f"=C{stats_row}-B{stats_row}"
    stats_row += 1

    # Mediana
    ws_comp[f'A{stats_row}'] = "Mediana"
    ws_comp[f'B{stats_row}'] = f"=MEDIAN('📊 Datos Originales'!I:I)"
    ws_comp[f'C{stats_row}'] = f"=MEDIAN('🎯 Scores Simulados'!P:P)"
    ws_comp[f'D{stats_row}'] = f"=C{stats_row}-B{stats_row}"
    stats_row += 1

    # Máximo
    ws_comp[f'A{stats_row}'] = "Máximo"
    ws_comp[f'B{stats_row}'] = f"=MAX('📊 Datos Originales'!I:I)"
    ws_comp[f'C{stats_row}'] = f"=MAX('🎯 Scores Simulados'!P:P)"
    ws_comp[f'D{stats_row}'] = f"=C{stats_row}-B{stats_row}"
    stats_row += 1

    # Mínimo
    ws_comp[f'A{stats_row}'] = "Mínimo"
    ws_comp[f'B{stats_row}'] = f"=MIN('📊 Datos Originales'!I:I)"
    ws_comp[f'C{stats_row}'] = f"=MIN('🎯 Scores Simulados'!P:P)"
    ws_comp[f'D{stats_row}'] = f"=C{stats_row}-B{stats_row}"
    stats_row += 1

    # Desviación estándar
    ws_comp[f'A{stats_row}'] = "Desviación Estándar"
    ws_comp[f'B{stats_row}'] = f"=STDEV('📊 Datos Originales'!I:I)"
    ws_comp[f'C{stats_row}'] = f"=STDEV('🎯 Scores Simulados'!P:P)"
    ws_comp[f'D{stats_row}'] = f"=C{stats_row}-B{stats_row}"

    ws_comp.column_dimensions['A'].width = 25
    ws_comp.column_dimensions['B'].width = 15
    ws_comp.column_dimensions['C'].width = 15
    ws_comp.column_dimensions['D'].width = 15

    # Guardar
    logger.info(f"\n3. Guardando Excel...")
    wb.save(OUTPUT_FILE)

    logger.info("\n" + "="*80)
    logger.info("✓ SIMULADOR CREADO EXITOSAMENTE")
    logger.info("="*80)
    logger.info(f"\n📁 Archivo: {OUTPUT_FILE.name}")
    logger.info(f"   Tamaño: {OUTPUT_FILE.stat().st_size / 1024:.1f} KB")
    logger.info(f"   Clientes: {len(df_sample)}")

    logger.info("\n📋 HOJAS DEL EXCEL:")
    logger.info("   1. 🎛️ PANEL DE CONTROL - Ajusta parámetros aquí")
    logger.info("   2. 📊 Datos Originales - Datos base de clientes")
    logger.info("   3. 🎯 Scores Simulados - Scores recalculados con tus parámetros")
    logger.info("   4. 📈 Comparación - Estadísticas base vs simulado")

    logger.info("\n🎯 CÓMO USAR:")
    logger.info("   1. Abre el Excel")
    logger.info("   2. Ve a la hoja '🎛️ PANEL DE CONTROL'")
    logger.info("   3. Cambia los valores en las celdas AZULES")
    logger.info("   4. Los scores se recalcularán automáticamente")
    logger.info("   5. Ve a '🎯 Scores Simulados' para ver resultados")
    logger.info("   6. Ve a '📈 Comparación' para ver impacto")

    logger.info("\n⚙️ PARÁMETROS AJUSTABLES:")
    logger.info("   • Pesos de componentes (Payment Perf, Plan, Deterioration)")
    logger.info("   • Penalizaciones (DPD >30d, Utilización >80%)")
    logger.info("   • Bonificaciones (>90% a tiempo, >24 meses)")
    logger.info("   • Pesos híbridos por categoría de madurez")
    logger.info("   • Umbrales de ratings (A+, A, B+, etc.)")

    logger.info("\n💡 SIMULACIONES SUGERIDAS:")
    logger.info("   1. Aumentar peso de Payment Performance a 700 pts")
    logger.info("   2. Penalizar más la alta utilización (30 pts)")
    logger.info("   3. Dar más peso a PLATAM para clientes maduros (80%)")
    logger.info("   4. Ajustar umbral de A+ de 900 a 920 pts")

    logger.info("\n" + "="*80)

    return OUTPUT_FILE

if __name__ == '__main__':
    create_interactive_simulator()
